import os
import openpyxl
from PyPDF2 import PdfReader
from collections import Counter
import re

def save_row(identity, df, out_path, resume_from=None):
    """Save data to Excel file."""
    file_name = out_path + '.xlsx'
    print("Saving to " + file_name)
    
    # Create new workbook if it doesn't exist
    if not os.path.exists(file_name):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = identity
        workbook.save(file_name)
        workbook.close()
    
    workbook = openpyxl.load_workbook(filename=file_name)
    sheet_name = identity
    
    # Create new sheet if it doesn't exist
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=sheet_name)
    else:
        sheet = workbook[sheet_name]
    
    print(f"\nSaving Outputs to Sheet {sheet_name}\n")
    
    if resume_from is not None and resume_from > 0:
        # If resuming, find the row where we need to start overwriting
        row_to_start = None
        for row in range(1, sheet.max_row + 1):
            if sheet_name == 'screening_results_summary':
                paper_num_cell = sheet.cell(row=row, column=1).value
                if isinstance(paper_num_cell, (int, float)) and int(paper_num_cell) >= resume_from:
                    row_to_start = row
                    break
        
        if row_to_start:
            # Delete all rows from resume point onwards
            sheet.delete_rows(row_to_start, sheet.max_row - row_to_start + 1)
    
    # Append the new data
    sheet.append(df)
    
    workbook.save(file_name)
    workbook.close()

def extract_text_from_pdf(pdf_path):
    """Extract text from PDF file."""
    reader = PdfReader(pdf_path)
    text = []
    
    for page in reader.pages:
        content = page.extract_text()
        if content:
            # Remove headers and footers (assuming they're in the first and last lines)
            lines = content.split('\n')
            if len(lines) > 2:
                content = '\n'.join(lines[1:-1])
            text.append(content)
    
    return ' '.join(text)

def clean_text(text):
    """Clean extracted text by removing common headers and splitting at references."""
    # Split at references section
    ref_patterns = r'(?<=\n)(References|Bibliography|Works Cited)'
    splits = re.split(ref_patterns, text, flags=re.IGNORECASE)
    main_text = splits[0] if splits else text
    
    # Remove common headers
    lines = main_text.split('\n')
    cleaned_lines = []
    
    # Count line frequencies to identify headers
    line_counter = Counter(lines)
    headers = {line for line, count in line_counter.items() if count >= 3 and len(line.split()) > 1}
    
    for line in lines:
        if line not in headers:
            cleaned_lines.append(line)
    
    return '\n'.join(cleaned_lines)

def full_text(paper_path):
    """Extract and clean text from PDF."""
    raw_text = extract_text_from_pdf(paper_path)
    cleaned_text = clean_text(raw_text)
    return cleaned_text
