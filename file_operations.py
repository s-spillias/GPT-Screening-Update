import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from auxiliary import save_row

def update_html(out_path, paper_num, title, summary_decision, update_screening_progress, n_studies):
    """Update the HTML file with the latest paper information"""
    html_path = os.path.join(out_path, 'screening_progress.html')
    
    # Ensure the directory exists
    os.makedirs(os.path.dirname(html_path), exist_ok=True)
    
    if not os.path.exists(html_path):
        # Create a new HTML file if it doesn't exist
        with open(html_path, 'w') as f:
            f.write('''
            <html>
            <head>
                <title>Screening Progress</title>
                <style>
                    table { border-collapse: collapse; width: 100%; }
                    th, td { border: 1px solid black; padding: 8px; text-align: left; }
                    th { background-color: #f2f2f2; }
                </style>
            </head>
            <body>
                <h1>Screening Progress</h1>
                <table>
                    <tr>
                        <th>Paper Number</th>
                        <th>Title</th>
                        <th>Decision</th>
                    </tr>
            ''')
    
    # Append the new paper information
    with open(html_path, 'a') as f:
        f.write(f'''
                <tr>
                    <td>{paper_num}</td>
                    <td>{title}</td>
                    <td>{summary_decision}</td>
                </tr>
        ''')
    
    # Close the HTML tags if it's the last paper
    if paper_num == n_studies - 1:
        with open(html_path, 'a') as f:
            f.write('''
                </table>
            </body>
            </html>
            ''')
    
    # Update the screening progress
    update_screening_progress(paper_num, title, summary_decision)

def get_last_processed_paper(model_to_use):
    """Get the last processed paper number from existing output"""
    output_dir = os.path.join('AI_Output', model_to_use)
    excel_file = os.path.join(output_dir, 'screening_results.xlsx')
    
    if os.path.exists(excel_file):
        try:
            df = pd.read_excel(excel_file, sheet_name='screening_results_summary')
            if not df.empty:
                # Ensure the first column is treated as numeric, ignoring non-numeric values
                df.iloc[:, 0] = pd.to_numeric(df.iloc[:, 0], errors='coerce')
                return int(df.iloc[:, 0].max())
        except Exception as e:
            print(f"Error reading Excel file: {str(e)}")
    return -1

def is_valid_excel(file_path):
    try:
        pd.read_excel(file_path)
        return True
    except:
        return False

def append_rows_to_excel(file_path, sheet_name, rows_data, headers=None):
    """Append multiple rows to an existing Excel sheet or create a new one"""
    try:
        print(f"Attempting to append rows to {file_path}, sheet: {sheet_name}")
        
        if not os.path.exists(file_path) or not is_valid_excel(file_path):
            print(f"Creating new Excel file: {file_path}")
            workbook = Workbook()
            workbook.save(file_path)
        
        workbook = load_workbook(file_path)
        
        if sheet_name not in workbook.sheetnames:
            print(f"Creating new sheet: {sheet_name}")
            workbook.create_sheet(sheet_name)
        
        sheet = workbook[sheet_name]
        
        if sheet.max_row == 1:  # Sheet is empty, write headers
            for col, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col, value=header)
        
        # Append new rows
        for row_data in rows_data:
            sheet.append(row_data)
        
        workbook.save(file_path)
        print(f"Successfully appended rows to {sheet_name} in {file_path}")
        
        # Verify that the file was updated
        if os.path.exists(file_path):
            print(f"File exists after save operation: {file_path}")
            print(f"File size: {os.path.getsize(file_path)} bytes")
        else:
            print(f"Error: File does not exist after save operation: {file_path}")
    except Exception as e:
        print(f"Error appending rows to Excel: {str(e)}")

def save_results(screen_name, out_path, paper_num, title, abstract, summary_decision, resume_from, n_agents, info_all, save_stuff, screening_criteria):
    try:
        print(f"\nSaving results for paper {paper_num}")
        summary_row = [paper_num, title, abstract, summary_decision]
        
        # Ensure the output directory exists
        os.makedirs(out_path, exist_ok=True)
        
        # Append to summary sheet
        excel_path = os.path.join(out_path, 'screening_results.xlsx')
        print(f"Saving to Excel file: {excel_path}")
        summary_headers = ["Paper Number", "Title", "Abstract", "Summary Decision"]
        append_rows_to_excel(excel_path, f"{screen_name}_summary", [summary_row], summary_headers)
        
        # Load the workbook once for all agent data
        workbook = load_workbook(excel_path)
        
        for agent in range(len(info_all)):
            try:
                save_info = [title, abstract, paper_num]
                for SC in screening_criteria:
                    try:
                        initial = save_stuff.get(SC['type'], {}).get('Initial', ['no info'] * n_agents)[agent]
                        final = save_stuff.get(SC['type'], {}).get('Final', ['no info'] * n_agents)[agent]
                        assessment = save_stuff.get(SC['type'], {}).get('Assessment', ['no info'] * n_agents)[agent]
                        
                        # Convert to string if not already
                        initial = str(initial) if initial is not None else 'no info'
                        final = str(final) if final is not None else 'no info'
                        assessment = str(assessment) if assessment is not None else 'no info'
                        
                        save_info.extend([initial, final, assessment])
                        
                        # Ensure the column exists and is of type object before assigning
                        for col_suffix in ['_Initial', '_Final', '_Assessment']:
                            col_name = f"{SC['type']}{col_suffix}"
                            if col_name not in info_all[agent].columns:
                                info_all[agent][col_name] = 'no info'
                            info_all[agent][col_name] = info_all[agent][col_name].astype(object)
                        
                        # Update info_all with string values
                        info_all[agent].at[paper_num, f"{SC['type']}_Initial"] = initial
                        info_all[agent].at[paper_num, f"{SC['type']}_Final"] = final
                        info_all[agent].at[paper_num, f"{SC['type']}_Assessment"] = assessment
                        
                    except Exception as e:
                        print(f"Error extending save_info for agent {agent}, criterion {SC['type']}: {str(e)}")
                        save_info.extend(['no info', 'no info', 'no info'])
                
                # Append to agent sheet
                agent_sheet_name = f"Agent_{agent}"
                print(f"Saving data for agent {agent} to sheet {agent_sheet_name}")
                
                if agent_sheet_name not in workbook.sheetnames:
                    print(f"Creating new sheet for Agent {agent}")
                    workbook.create_sheet(agent_sheet_name)
                
                sheet = workbook[agent_sheet_name]
                
                agent_headers = ["Title", "Abstract", "Paper Number"] + [f"{SC['type']}_{suffix}" for SC in screening_criteria for suffix in ['Initial', 'Final', 'Assessment']]
                
                if sheet.max_row == 1:  # Sheet is empty, write headers
                    sheet.append(agent_headers)
                
                sheet.append(save_info)
                print(f"Appended data for paper {paper_num} to Agent {agent}'s sheet")
                
            except Exception as e:
                print(f"Could not save data for agent {agent}: {str(e)}")
        
        # Save the workbook after all agents' data has been written
        workbook.save(excel_path)
        print(f"Saved all data to {excel_path}")
        
        # Verify that the file exists and has been updated
        if os.path.exists(excel_path):
            print(f"Excel file exists after save operation: {excel_path}")
            print(f"File size: {os.path.getsize(excel_path)} bytes")
            print(f"Last modified: {os.path.getmtime(excel_path)}")
            
            # Verify each agent's sheet
            workbook = load_workbook(excel_path)
            for agent in range(len(info_all)):
                agent_sheet_name = f"Agent_{agent}"
                if agent_sheet_name in workbook.sheetnames:
                    sheet = workbook[agent_sheet_name]
                    print(f"Agent {agent}'s sheet exists with {sheet.max_row} rows")
                else:
                    print(f"Error: Sheet for Agent {agent} not found")
        else:
            print(f"Error: Excel file does not exist after save operation: {excel_path}")
    
    except Exception as e:
        print(f"Error in save_results function: {str(e)}")
